import openpyxl  # type: ignore[import]
import requests
from pathlib import Path
from typing import Dict, List

apocalyptechSheet = "https://docs.google.com/spreadsheets/d/1v-F_3C2ceaFKJae1b6wmbelw_jLjmPPriBLzGTZMqRc/export?format=xlsx&id=1v-F_3C2ceaFKJae1b6wmbelw_jLjmPPriBLzGTZMqRc"
characters = {
    "FL4K": "Beastmaster",
    "Moze": "Gunner",
    "Zane": "Operative",
    "Amara": "Siren",
}

emotes, decos, heads, skins, echos, trinkets, weaponSkins = [], [], [], [], [], [], []

playerHeads: Dict[str, List[str]] = {x: [] for x, _ in characters.items()}
playerSkins: Dict[str, List[str]] = {x: [] for x, _ in characters.items()}


def requestGoogleSheetAsXLSX(url):
    print("Requesting google sheet: {}".format(url))
    r = requests.get(url, allow_redirects=True)
    open("sheetData.xlsx", "wb").write(r.content)
    print("Google sheet written...")


requestGoogleSheetAsXLSX(apocalyptechSheet)

xlsxFile = Path("sheetData.xlsx")
workbook = openpyxl.load_workbook(xlsxFile)
sheetNames = workbook.sheetnames[1:]
print("Sheet Names: {}".format(sheetNames))

for sheetName in sheetNames:
    worksheet = workbook[sheetName]

    for row in range(
        2, worksheet.max_row + 1
    ):  # Iterate through each row of the worksheet
        invBalance = worksheet["C{}".format(row)].value
        customizationName = worksheet["A{}".format(row)].value

        if invBalance == None:
            break

        assetPath = invBalance.replace("InvBal_", "")

        if "Heads" in sheetName:
            heads.append((customizationName, assetPath))
            print(
                f"\t[.] Parsing heads: {customizationName} on sheet: {sheetName} :: Asset Path: {assetPath}"
            )
            playerHeads[[x for x, y in characters.items() if y in sheetName][0]].append(
                assetPath
            )
        elif "Skins" in sheetName and "Weapon" not in sheetName:
            skins.append((customizationName, assetPath))
            print(f"\t[..] Parsing skins: {customizationName} on sheet: {sheetName}")
            playerSkins[[x for x, y in characters.items() if y in sheetName][0]].append(
                assetPath
            )
        elif "Emotes" in sheetName:
            emotes.append((customizationName, assetPath))
        elif "Themes" in sheetName:
            echos.append((customizationName, assetPath))
        elif "Decorations" in sheetName:
            decos.append((customizationName, assetPath))
        elif "Trinket" in sheetName:
            trinkets.append((customizationName, assetPath))
        elif "Weapon Skins" in sheetName:
            weaponSkins.append((customizationName, assetPath))

    print("Done reading {}".format(sheetName))

pathFormat = "public static readonly Dictionary<string, string> {0}AssetPaths = new Dictionary<string, string>()"

assetSets = []
nameToList = {
    "emotes": emotes,
    "deco": decos,
    "head": heads,
    "skin": skins,
    "echo": echos,
    "weapon": weaponSkins,
    "trinket": trinkets,
}

for name in nameToList:
    list = nameToList[name]
    assetSet = pathFormat.format(name) + "{\n"
    for name, asset in list:
        assetSet += '            {{"{0}", "{1}"}},\n'.format(asset, name)
    assetSet = assetSet[:-2] + "\n        };\n"
    assetSets += [assetSet]

pathFormat = "public static readonly Dictionary<string, List<string>> {0}NamesDictionary = new Dictionary<string, List<string>>()"

assetSet = pathFormat.format("Head") + "{\n"
for char, hx in playerHeads.items():
    assetSet += '            {{"{0}", new List<string>() {{ {1} }} }},\n'.format(
        char, str(hx)[1:-1].replace("'", '"')
    )
assetSet = assetSet[:-2] + "\n        };\n"
assetSets += [assetSet]

assetSet = pathFormat.format("Skin") + "{\n"
for char, hx in playerSkins.items():
    assetSet += '            {{"{0}", new List<string>() {{ {1} }} }},\n'.format(
        char, str(hx)[1:-1].replace("'", '"')
    )
assetSet = assetSet[:-2] + "\n        };\n"
assetSets += [assetSet]

with open("output.txt", "w") as outFile:
    completeSet = ""

    for assetSet in assetSets:
        completeSet += assetSet + "\n"

    outFile.write(completeSet)
